# -*- coding: utf-8 -*-
import requests
import json
import openpyxl
from tkinter import *
from tkinter.filedialog import askdirectory
import os
import threading
from time import time, sleep
import io


def write_json_notebook(data, filename='json_notebook'):
    with open(filename, 'w', encoding='utf8') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
    file.close()


def write_json_motherboard(data, filename='json_motherboard'):
    with open(filename, 'w', encoding='utf8') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
    file.close()


def write_json_category1(data, filename='category_brain'):
    with open(filename, 'w') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
    file.close()

#
# def get_sid() -> str:
#     data = {'login': 'test_123', 'password': '6b866754d9c9aa72dd660e5ff5491b2b'}
#     r_sid = requests.post('http://api.brain.com.ua/auth', data)
#     json_data_sid = r_sid.json()
#     sid = json_data_sid['result']
#     return sid
#
#
# def category_motherboard():
#     r_category = requests.get('http://api.brain.com.ua/products/1264/' + get_sid())
#     json_data_category_motherboard = r_category.json()
#     write_json_category1(json_data_category_motherboard)
#     return json_data_category_motherboard
# category_motherboard()
#
def content_motherboard():
    json_data_content = category_motherboard()
    product_list_id = [productIDs['productID'] for productIDs in json_data_content['result']['list']]
    product_list_motherboard = []
    for productID in product_list_id:
        r_content = requests.post('http://api.brain.com.ua/products/content/' + get_sid(), params={'productIDs': productID})
        json_data = r_content.json()
        product_list_motherboard.append(json_data)
        write_json_motherboard(product_list_motherboard)
#     return product_list_motherboard


def set_header_motherboard():
    json_motherboard = content_motherboard()
    count = 0
    for i in json_motherboard[0]['result']['options']:
        count += 1
    return count * 2


def write_xlsx_motherboard():
    # json_content = content_motherboard()
    # json_category = category_motherboard()
    with open('json_motherboard') as file:
        json_content = json.load(file)
    with open('category_brain') as file:
        json_category = json.load(file)
    header = [
        'PRODUCT_ID', 'NAME', 'ARTICUL', 'PRODUCT_CODE',
        'RETAIL_PRICE_UAH', 'AVAILABLE',
        'BRIEF_DESCRIPTION', 'MEDIUM_IMAGE', 'DESCRIPTION'
    ]
    set_json = set_header_motherboard()
    for i in range(set_json):
        if i % 2 == 0:
            header.append('OPTION_NAME')
        else:
            header.append('VALUE_NAME')

    book = openpyxl.Workbook()
    sheet = book.active
    j = 1
    for val in header:
        sheet.cell(row=1, column=j).value = val
        j += 1
    row = 2
    for product in json_category['result']['list']:
        sheet[row][0].value = product['productID']
        sheet[row][1].value = product['name']
        sheet[row][2].value = product['articul']
        sheet[row][3].value = product['product_code']
        sheet[row][4].value = product['retail_price_uah']
        sheet[row][5].value = str(product['available'])
        sheet[row][6].value = product['brief_description']
        sheet[row][7].value = product['medium_image']
        #sheet[row][8].value = product['description']
        row += 1
    row = 2
    column = 9
    for option in json_content[0]['result']['options'][0]:
        sheet.cell(row=row, column=column).value = option['name']
        column += 1
        sheet.cell(row=row, column=column).value = option['value']
        column += 1
    row += 1


    # dirname = askdirectory(initialdir=os.getcwd(), title='Please select a directory')
    # dirname = book.save(str(dirname) + '/motherboard_brain.xlsx')
    book.save('_brain.xlsx')
    book.close()
write_xlsx_motherboard()





# def start_motherboard():
#     threading.Thread(target=write_xlsx_motherboard).start()
#
#
# def start_cpu():
#     threading.Thread(target=write_xlsx_cpu).start()
#
#
# root = Tk()
# root.title('API BRAIN')
# root.geometry("300x200")
# root.resizable(False, False)
# # # path = StringVar()
# # # path_entry = Entry(root,textvariable=path, background="#ccc", width=45)
# # # path_entry.place(x=10, y=20)
# # # path_entry.insert(0, "http://api.brain.com.ua/products/")
# c1 = Checkbutton(root, text="Материнские платы", command=start_motherboard)
# c1.pack(anchor=W, padx=10)
# c2 = Checkbutton(root, text="Процессоры", command=start_cpu)
# c2.pack(anchor=W, padx=10)
# # # btn = Button(root, text="Send", background="#555", foreground="#ccc", padx="30", pady="1", font="8", command=start)
# # # btn.place(x=110, y=110)
# root.mainloop()
