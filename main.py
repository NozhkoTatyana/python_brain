# -*- coding: utf-8 -*-
import requests
import json
import openpyxl
from tkinter import *
from tkinter.filedialog import askdirectory
import os
import threading
from time import time, sleep
from woocommerce import *
import win32com.client as win32
import pythoncom
import pandas as pd
import  io




def write_json_notebook(data, filename='json_notebook'):
    with open(filename, 'w', encoding='utf8') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
    file.close()

def write_json_motherboard(data, filename='json_motherboard'):
    with open(filename, 'w',  encoding='utf8') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
    file.close()

def write_json_category1(data, filename='category_brain'):
    with open(filename, 'w') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
    file.close()



def get_sid() -> str:
    data = {'login': 'test_123', 'password': '6b866754d9c9aa72dd660e5ff5491b2b'}
    r_sid = requests.post('http://api.brain.com.ua/auth', data)
    json_data_sid = r_sid.json()
    sid = json_data_sid['result']
    return sid

# def  get_gategories_id():
#     r_id = requests.get('http://api.brain.com.ua/categories/' + get_sid())
#     data_id = r_id.json()
#     return  data_id
#
# def get_category():
#     id = get_gategories_id()
#     list_id = [categories_id['categoryID'] for categories_id in id['result']]
#     id_category = []
#     for categories_id in list_id:
#         r_category = requests.get('http://api.brain.com.ua/products/' + str(categories_id) + '/' + get_sid())
#         json_data_category = r_category.json()
#         id_category.append(json_data_category)
#         write_json_category(id_category)
#         return id_category
#     #write_json_category(json_data_category)

def category_motherboard():
    r_category = requests.get('http://api.brain.com.ua/products/1264/' + get_sid())
    json_data_category_motherboard = r_category.json()
    return json_data_category_motherboard


def content_motherboard():
    json_data_content = category_motherboard()
    product_list_code = [product_code['product_code'] for product_code in json_data_content['result']['list']]
    product_list_motherboard = []
    for product_code in product_list_code:
        r_content = requests.get('http://api.brain.com.ua/product/product_code/' + product_code + '/' + get_sid())
        json_data = r_content.json()
        product_list_motherboard.append(json_data)
        #return product_list_motherboard
        write_json_motherboard(product_list_motherboard)
        sleep(1)

    data_motherboard = json.load(io.open('json_motherboard', encoding='utf-8').read())
    # data_motherboard = content_motherboard()

    pythoncom.CoInitializeEx(0)
    ExcelApp = win32.Dispatch('Excel.Application')
    ExcelApp.Visible = True
    wb = ExcelApp.Workbooks.Add()
    ws = wb.Worksheets(1)
    rows = []
    for record in data_motherboard:
        productID = record['result']['productID']
        name = record['result']['name']
        articul = record['result']['articul']
        product_code = record['result']['product_code']
        retail_price_uah = record['result']['retail_price_uah']
        available = str(record['result']['available'])
        brief_description = record['result']['brief_description']
        country = record['result']['country']
        medium_image = record['result']['medium_image']
        description = record['result']['description']
        warranty = record['result']['warranty']
        options = ';'.join(
            ['{0}, {1}'.format(option['name'], option['value']) for option in record['result']['options']])

        rows.append([
            productID, name,
            articul, product_code,
            retail_price_uah, available,
            brief_description, country,
            medium_image, description, warranty,
            options
        ])

    header_labels = ('PRODUCT_ID', 'NAME', 'ARTICUL', 'PRODUCT_CODE',
                     'RETAIL_PRICE_UAH', 'AVAILABLE',
                     'BRIEF_DESCRIPTION', 'COUNTRY', 'MEDIUM_IMAGE', 'DESCRIPTION', 'WARRANTY', 'OPTIONS')
    for index, val in enumerate(header_labels):
        ws.Cells(1, index + 1).Value = val
        row_tracker = 2
        column_size = len(header_labels)
    for row in rows:
        ws.Range(
            ws.Cells(row_tracker, 1),
            ws.Cells(row_tracker, column_size)
        ).value = row
        row_tracker += 1
    wb.SaveAs(os.path.join(os.getcwd(), 'motherboard.xlsx'))
    ExcelApp.Quit()
    ExcelApp = None

def category_notebook():
    r_category = requests.get('http://api.brain.com.ua/products/1191/' + get_sid())
    json_data_category_notebook = r_category.json()
    return json_data_category_notebook


def content_notebook():
    json_data_content = category_notebook()
    product_list_code = [product_code['product_code'] for product_code in json_data_content['result']['list']]
    product_list_notebook = []
    for product_code in product_list_code:
        r_content = requests.get('http://api.brain.com.ua/product/product_code/' + product_code + '/' + get_sid())
        json_data = r_content.json()
        product_list_notebook.append(json_data)
        #return product_list_notebook
        #write_json_notebook(product_list_notebook)
        sleep(1)

        #data_notebook = json.loads(io.open('json_notebook', encoding='utf-8').read())
        #data_notebook = content_notebook()
    pythoncom.CoInitializeEx(0)
    ExcelApp = win32.Dispatch('Excel.Application')
    ExcelApp.Visible = True
    wb = ExcelApp.Workbooks.Add()
    ws = wb.Worksheets(1)
    rows = []
    for record in product_list_notebook:
            productID = record['result']['productID']
            name = record['result']['name']
            articul = record['result']['articul']
            product_code = record['result']['product_code']
            retail_price_uah = record['result']['retail_price_uah']
            available = str(record['result']['available'])
            brief_description = record['result']['brief_description']
            country = record['result']['country']
            medium_image = record['result']['medium_image']
            description = record['result']['description']
            warranty = record['result']['warranty']
            options = ';'.join(['{0}, {1}'.format(option['name'],option['value']) for option in record['result']['options']])

            rows.append([
                productID, name,
                articul, product_code,
                retail_price_uah, available,
                brief_description, country,
                medium_image, description, warranty,
                options
            ])

    header_labels = ('PRODUCT_ID', 'NAME', 'ARTICUL', 'PRODUCT_CODE',
        'RETAIL_PRICE_UAH', 'AVAILABLE',
        'BRIEF_DESCRIPTION', 'COUNTRY', 'MEDIUM_IMAGE', 'DESCRIPTION', 'WARRANTY', 'OPTIONS')
    for index, val in enumerate(header_labels):
           ws.Cells(1, index + 1).Value = val
           row_tracker = 2
           column_size = len(header_labels)
    for row in rows:
            ws.Range(
                ws.Cells(row_tracker, 1),
                ws.Cells(row_tracker, column_size)
            ).value = row
            row_tracker += 1
    wb.SaveAs(os.path.join(os.getcwd(), 'notebook.xlsx'))
    ExcelApp.Quit()
    ExcelApp = None

# def set_header():
#     json_data_content = show_computer()
#     count = 0
#     for content in json_data_content[0]['result']['list'][0]['options']:
#         count += 1
#     return count*2

#
# def write_xlsx():
#     #json_data_category_1 = get_category()
#     #json_content = get_content()
#     product_list = json.loads(io.open('content_brain', encoding='utf-8').read())
#     header = [
#         'PRODUCT_ID', 'NAME', 'ARTICUL', 'PRODUCT_CODE',
#         'RETAIL_PRICE_UAH', 'STOCKS', 'STOCKS_EXPECTED', 'AVAILABLE',
#         'BRIEF_DESCRIPTION', 'COUNTRY', 'MEDIUM_IMAGE', 'WARRANTY',
#         'MODEL', 'DESCRIPTION', 'WEIGHT', 'DATE_ADDED','OPTION_NAME', 'VALUE_NAME'
#     ]
#     # set_json = set_header()
#     # for i in range(set_json):
#     #     if i % 2 == 0:
#     #         header.append('OPTION_NAME')
#     #     else:
#     #         header.append('VALUE_NAME')
#
    # book = openpyxl.Workbook()
    # sheet = book.active
#     # j = 1
#     # for val in header:
#     #     sheet.cell(row=1, column=j).value = val
#     #     j += 1
#     row = 2
#     for product in product_list:
#         print(product['result']['name'])
            # sheet[row][0].value = product['productID']
            # sheet[row][1].value = product['name']
            # sheet[row][2].value = product['articul']
            # sheet[row][3].value = product['product_code']
            # sheet[row][4].value = product['retail_price_uah']
            # sheet[row][5].value = str(product['stocks'])
            # sheet[row][6].value = str(product['stocks_expected'])
            # sheet[row][7].value = str(product['available'])
            # sheet[row][8].value = product['brief_description']
            # sheet[row][9].value = product['country']
            # sheet[row][10].value = product['medium_image']
            # sheet[row][11].value = product['warranty']
            # row += 1
#     # row = 2
#     # column = 16
#     # for content in json_content:
#     #     del content['status']
#     #     for content_list in content['result']['list']:
#     #         sheet[row][12].value = content_list['model']
#     #         sheet[row][13].value = content_list['description']
#     #         sheet[row][14].value = content_list['weight']
#     #         sheet[row][15].value = content_list['date_added']
#     #         for option in content_list['options']:
#     #             sheet.cell(row=row, column=column).value = option['OptionName']
#     #             column += 1
#     #             sheet.cell(row=row, column=column).value = option['ValueName']
#     #             column += 1
#     #         row += 1
#     #         column = 16
#
#     # dirname = askdirectory(initialdir=os.getcwd(), title='Please select a directory')
#     # dirname = book.save(str(dirname) + '/_brain.xlsx')
#     book.save('_brain.xlsx')
#     book.close()
# write_xlsx()


def start_motherboard():
    threading.Thread(target=content_motherboard).start()
def start_notebook():
    threading.Thread(target=content_notebook).start()


root = Tk()
root.title('API BRAIN')
root.geometry("300x200")
root.resizable(False, False)
# # path = StringVar()
# # path_entry = Entry(root,textvariable=path, background="#ccc", width=45)
# # path_entry.place(x=10, y=20)
# # path_entry.insert(0, "http://api.brain.com.ua/products/")
c1 = Checkbutton(root, text="Материнские платы", command=start_motherboard)
c1.pack(anchor=W, padx=10)
c2 = Checkbutton(root, text="Ноутбуки", command=start_notebook)
c2.pack(anchor=W, padx=10)
# # btn = Button(root, text="Send", background="#555", foreground="#ccc", padx="30", pady="1", font="8", command=start)
# # btn.place(x=110, y=110)
root.mainloop()




