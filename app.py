# -*- coding: utf-8 -*-
import json, codecs
import requests
import openpyxl
import xlrd
import csv
import pandas as pd
from time import sleep
from datetime import datetime
from openpyxl import load_workbook



def get_sid() -> str:
    data = {'login': 'test_123', 'password': 'test_123'}
    r_sid = requests.post('http://api.brain.com.ua/auth', data)
    json_data_sid = r_sid.json()
    sid = json_data_sid['result']
    return sid


def category_notebook():
    # print('---Список id категорий---')
    # print('1097 ---> Процессоры')
    # print('1264 ---> Материнские платы')
    # print('1191 ---> Ноутбуки')
    # test_text = input("Введите id категории:")
    print("Получил SID, приступаю к работе")
    r_category = requests.get('http://api.brain.com.ua/products/1191/' + get_sid())  # + test_text + '/'.
    print("Приступаю к обработке категории ноутбуков")
    json_data_category_notebook = r_category.json()
    with open('category_notebook.json', 'wb') as f:
        json.dump(json_data_category_notebook, codecs.getwriter('utf-8')(f), ensure_ascii=False)
#sleep(5)
category_notebook()


def content_notebook():
    with open('category_notebook.json') as f:
        json_data_content = json.load(f)
    product_list_code = [product_code['product_code'] for product_code in json_data_content['result']['list']]
    product_list_notebook = []
    for product_code in product_list_code:
        r_content = requests.get('http://api.brain.com.ua/product/product_code/' + product_code + '/' + get_sid())
        json_data = r_content.json()
        product_list_notebook.append(json_data)
        with open('data_content_notebook.json', 'wb') as f:
            json.dump(product_list_notebook, codecs.getwriter('utf-8')(f), ensure_ascii=False)
        print("Записываю джейсон notebook")
content_notebook()




def set_header_notebook():
    #json_notebook = content_notebook()
    with open('data_content_notebook.json', encoding='utf8') as file:
        json_notebook = json.load(file)
    count = 0
    for i in json_notebook[0]['result']['options']:
        count += 1
    return count * 2


def write_xlsx_notebook():
    print('Начинаю собирать файл')
    #json_content = content_notebook()
    # json_category = category_motherboard()
    with open('data_content_notebook.json', encoding='utf8') as file:
        json_content = json.load(file)
    file.close()
    header = [
        'PRODUCT_CODE', 'NAME', 'BRIEF_DESCRIPTION',
        'RETAIL_PRICE_UAH', 'DESCRIPTION', 'AVAILABLE',
        'MEDIUM_IMAGE', 'CATEGORY', 'STOCKS_EXPECTED', 'OPTION_NAME', 'VALUE_NAME'
    ]
    set_json_n = set_header_notebook()
    for i in range(set_json_n):
        if i % 2 == 0:
            header.append('OPTION_NAME')
        else:
            header.append('VALUE_NAME')

    book = openpyxl.Workbook()
    sheet = book.active
    j = 1
    for val in header:
        sheet.cell(row=1, column=j).value = val
        j += 1
    row = 2
    for product in json_content:
        if len(product['result']['available']) != 0 or len(product['result']["stocks_expected"]) != 0:
            sheet[row][0].value = product['result']['product_code']
            sheet[row][1].value = product['result']['name']
            sheet[row][2].value = product['result']['brief_description']
            sheet[row][3].value = product['result']['retail_price_uah']
            sheet[row][4].value = product['result']['description']
            sheet[row][5].value = str(product['result']['available'])
            sheet[row][6].value = product['result']['medium_image']
            sheet[row][7].value = str('Ноутбуки')
            sheet[row][8].value = str(product['result']["stocks_expected"])
            row += 1
    row = 2
    column = 10
    for option in json_content:
        if len(option['result']['available']) != 0 or len(option['result']["stocks_expected"]) != 0:
            for s in option['result']['options']:
                sheet.cell(row=row, column=column).value = s['name']
                column += 1
                sheet.cell(row=row, column=column).value = s['value']
                column += 1
            row += 1
            column = 10

    # dirname = askdirectory(initialdir=os.getcwd(), title='Please select a directory')
    # dirname = book.save(str(dirname) + '/_brain.xlsx')
    book.save('_brain.xlsx')
    print("Загрузка категории ноутбуков завершена!")
write_xlsx_notebook()

def category_notebook():
    # print('---Список id категорий---')
    # print('1097 ---> Процессоры')
    # print('1264 ---> Материнские платы')
    # print('1191 ---> Ноутбуки')
    # test_text = input("Введите id категории:")
    r_category = requests.get('http://api.brain.com.ua/products/1264/' + get_sid())  # + test_text + '/'.
    print("Приступаю к обработке категории материнских плат")
    json_data_category_notebook = r_category.json()
    with open('category_motherboard.json', 'wb') as f:
        json.dump(json_data_category_notebook, codecs.getwriter('utf-8')(f), ensure_ascii=False)
#sleep(5)
category_notebook()


def content_notebook():
    with open('category_motherboard.json') as f:
        json_data_content = json.load(f)
    product_list_code = [product_code['product_code'] for product_code in json_data_content['result']['list']]
    product_list_notebook = []
    for product_code in product_list_code:
        r_content = requests.get('http://api.brain.com.ua/product/product_code/' + product_code + '/' + get_sid())
        json_data = r_content.json()
        product_list_notebook.append(json_data)
        with open('data_content_motherboard.json', 'wb') as f:
            json.dump(product_list_notebook, codecs.getwriter('utf-8')(f), ensure_ascii=False)
        print("Записываю джейсон материнских плат")
content_notebook()

def write_xlsx_notebook():
    print('Начинаю собирать файл')
    #json_content = content_notebook()
    # json_category = category_motherboard()
    with open('data_content_motherboard.json', encoding='utf8') as file:
        json_content = json.load(file)
    file.close()
    book = load_workbook('_brain.xlsx')
    sheet = book.active
    row = sheet.max_row
    row_option = row
    for product in json_content:
        if len(product['result']['available']) != 0 or len(product['result']["stocks_expected"]) != 0:
            sheet[row][0].value = product['result']['product_code']
            sheet[row][1].value = product['result']['name']
            sheet[row][2].value = product['result']['brief_description']
            sheet[row][3].value = product['result']['retail_price_uah']
            sheet[row][4].value = product['result']['description']
            sheet[row][5].value = str(product['result']['available'])
            sheet[row][6].value = product['result']['medium_image']
            sheet[row][7].value = str('Материнские платы')
            sheet[row][8].value = str(product['result']["stocks_expected"])
            row += 1
    row = row_option
    column = 10
    for option in json_content:
        if len(option['result']['available']) != 0 or len(option['result']["stocks_expected"]) != 0:
            for s in option['result']['options']:
                sheet.cell(row=row_option, column=column).value = s['name']
                column += 1
                sheet.cell(row=row_option, column=column).value = s['value']
                column += 1
            row_option += 1
            column = 10
    #print(column, row)
    # dirname = askdirectory(initialdir=os.getcwd(), title='Please select a directory')
    # dirname = book.save(str(dirname) + '/_brain.xlsx')
    book.save('_brain.xlsx')
    # book.close()
    # sheet = xlrd.open_workbook("_brain.xlsx").sheet_by_index(0)
    # col = csv.writer(open("_brain.csv", 'w', newline="", encoding='utf8'))
    # for row in range(sheet.nrows):
    #     col.writerow(sheet.row_values(row))
    # df = pd.DataFrame(pd.read_csv("_brain.csv"))
    print("Загрузка категории материнских плат завершена!")
write_xlsx_notebook()

def category_notebook():
    # print('---Список id категорий---')
    # print('1097 ---> Процессоры')
    # print('1264 ---> Материнские платы')
    # print('1191 ---> Ноутбуки')
    # test_text = input("Введите id категории:")
    r_category = requests.get('http://api.brain.com.ua/products/1097/' + get_sid())  # + test_text + '/'.
    print("Приступаю к следующей обработки категории")
    json_data_category_notebook = r_category.json()
    with open('category_cpu.json', 'wb') as f:
        json.dump(json_data_category_notebook, codecs.getwriter('utf-8')(f), ensure_ascii=False)
#sleep(5)
category_notebook()


def content_notebook():
    with open('category_cpu.json') as f:
        json_data_content = json.load(f)
    product_list_code = [product_code['product_code'] for product_code in json_data_content['result']['list']]
    product_list_notebook = []
    for product_code in product_list_code:
        r_content = requests.get('http://api.brain.com.ua/product/product_code/' + product_code + '/' + get_sid())
        json_data = r_content.json()
        product_list_notebook.append(json_data)
        with open('data_content_cpu.json', 'wb') as f:
            json.dump(product_list_notebook, codecs.getwriter('utf-8')(f), ensure_ascii=False)
        print("Записываю джейсон cpu")
content_notebook()


def write_xlsx_notebook():
    print('Начинаю собирать файл')
    #json_content = content_notebook()
    # json_category = category_motherboard()
    with open('data_content_cpu.json', encoding='utf8') as file:
        json_content = json.load(file)
    file.close()
    book = load_workbook('_brain.xlsx')
    sheet = book.active
    row = sheet.max_row
    row_option = row
    for product in json_content:
        if len(product['result']['available']) != 0 or len(product['result']["stocks_expected"]) != 0:
            sheet[row][0].value = product['result']['product_code']
            sheet[row][1].value = product['result']['name']
            sheet[row][2].value = product['result']['brief_description']
            sheet[row][3].value = product['result']['retail_price_uah']
            sheet[row][4].value = product['result']['description']
            sheet[row][5].value = str(product['result']['available'])
            sheet[row][6].value = product['result']['medium_image']
            sheet[row][7].value = str('Процессоры')
            sheet[row][8].value = str(product['result']["stocks_expected"])
            row += 1
    row = row_option
    column = 10
    for option in json_content:
        if len(option['result']['available']) != 0 or len(option['result']["stocks_expected"]) != 0:
            for s in option['result']['options']:
                sheet.cell(row=row_option, column=column).value = s['name']
                column += 1
                sheet.cell(row=row_option, column=column).value = s['value']
                column += 1
            row_option += 1
            column = 10

    # dirname = askdirectory(initialdir=os.getcwd(), title='Please select a directory')
    # dirname = book.save(str(dirname) + '/_brain.xlsx')
    book.save('_brain.xlsx')
    book.close()
    sheet = xlrd.open_workbook("_brain.xlsx").sheet_by_index(0)
    col = csv.writer(open("_brain.csv", 'w', newline="", encoding='utf8'))
    for row in range(sheet.nrows):
        col.writerow(sheet.row_values(row))
    df = pd.DataFrame(pd.read_csv("_brain.csv"))
    print("Загрузка завершена!")
write_xlsx_notebook()






