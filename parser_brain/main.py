import requests
import json
import openpyxl
import csv
from tkinter import *
from openpyxl import load_workbook
from tkinter import filedialog



def write_json(data, filename='cpu_brain.json'):
    with open(filename, 'w') as file:
        json.dump(data, file, indent=2, ensure_ascii=False)
    get_product()



def write_xlsx():
    with open('cpu_brain.json') as file:
        json_data = json.load(file)

        book = openpyxl.Workbook()
        sheet = book.active
        sheet['A1'] = 'NAME'
        sheet['B1'] = 'ARTICUL'
        sheet['C1'] = 'PRODUCT_CODE'
        sheet['D1'] = 'RETAIL_PRICE_UAH'
        sheet['E1'] = 'STOCKS'
        sheet['F1'] = 'AVAILABLE'
        sheet['G1'] = 'BRIEF_DESCRIPTION'
        sheet['H1'] = 'COUNTRY'
        sheet['I1'] = 'MEDIUM_IMAGE'
        sheet['J1'] = 'WARRANTY'

        row = 2

        for li in json_data['result']['list']:
            sheet[row][0].value = li['name']
            sheet[row][1].value = li['articul']
            sheet[row][2].value = li['product_code']
            sheet[row][3].value = li['retail_price_uah']
            #sheet[row][4].value = ' '.join(li['stocks'])
            sheet[row][5].value = ' '.join(li['available'])
            sheet[row][6].value = li['brief_description']
            sheet[row][7].value = li['country']
            sheet[row][8].value = li['medium_image']
            sheet[row][9].value = li['warranty']
            row += 1


        #save_spot = filedialog.askdirectory()
        book.save('_brain.xlsx')
        book.close()
        dir_path = filedialog.askdirectory()
        reader = load_workbook('_brain.xlsx')
        dir_path = filedialog.asksaveas('w')




def get_sid():
    data = {'login': 'test_123', 'password': '6b866754d9c9aa72dd660e5ff5491b2b'}
    r = requests.post('http://api.brain.com.ua/auth', data)
    jsonRes = r.json()
    sid = jsonRes['result']
    return sid

def get_product():
    url = 'http://api.brain.com.ua/products/1097/' + get_sid()
    r1 = requests.get(url)
    package_json = r1.json()



# #http://api.brain.com.ua/products/1097/
# root = Tk()
# root.title('API BRAIN')
# path = StringVar()
# entry = Entry(root, width=20, bg="#ebebeb", fg='#2c2c2c', textvariable=path)
# entry.place(relx=0.1, rely=0.3)
# btn = Button(root, text="Send", command=get_product())
# btn.place(relx=0.05, rely=0.05)
#
# root.mainloop()

